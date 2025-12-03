# Export Views - Excel and PDF
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from .models import Compra, OrdenCompra, Proveedor

# Import admin_required from views
def _admin_check(user):
    return user.is_authenticated and user.is_staff
admin_required = user_passes_test(_admin_check, login_url='login')

@admin_required
def export_compras_excel(request):
    """Exporta la lista de compras a Excel."""
    compras = Compra.objects.all().order_by('-fecha_compra')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Compras"
    
    # Header
    headers = ['Proveedor', 'Número Factura', 'Fecha Compra', 'Monto Total', 'Dirección', 'Correo']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for compra in compras:
        ws.append([
            compra.proveedor.nombre if compra.proveedor else compra.nombre_proveedor,
            compra.numero_factura,
            compra.fecha_compra.strftime('%d/%m/%Y'),
            float(compra.monto_total),
            compra.direccion_entrega,
            compra.correo_contacto
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_compras.xlsx'
    wb.save(response)
    return response

@admin_required
def export_compras_pdf(request):
    """Exporta la lista de compras a PDF."""
    compras = Compra.objects.all().order_by('-fecha_compra')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_compras.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Reporte de Compras</b>", styles['Title'])
    elements.append(title)
    
    # Table data
    data = [['Proveedor', 'Nº Factura', 'Fecha', 'Monto']]
    for compra in compras:
        data.append([
            compra.proveedor.nombre if compra.proveedor else compra.nombre_proveedor,
            compra.numero_factura,
            compra.fecha_compra.strftime('%d/%m/%Y'),
            f'${compra.monto_total:,.0f}'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@admin_required
def export_ordenes_excel(request):
    """Exporta la lista de órdenes de compra a Excel."""
    ordenes = OrdenCompra.objects.all().order_by('-fecha_emision')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes de Compra"
    
    # Header
    headers = ['ID Orden', 'Solicitante', 'Fecha Emisión', 'Neto', 'IVA', 'Total', 'Estado']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for orden in ordenes:
        ws.append([
            f'OC-{orden.id}',
            orden.solicitante.username if orden.solicitante else 'N/A',
            orden.fecha_emision.strftime('%d/%m/%Y'),
            float(orden.neto),
            float(orden.iva),
            float(orden.monto_total),
            orden.get_estado_display()
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ordenes_compra.xlsx'
    wb.save(response)
    return response

@admin_required
def export_ordenes_pdf(request):
    """Exporta la lista de órdenes de compra a PDF."""
    ordenes = OrdenCompra.objects.all().order_by('-fecha_emision')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=ordenes_compra.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Órdenes de Compra</b>", styles['Title'])
    elements.append(title)
    
    # Table data
    data = [['ID', 'Solicitante', 'Fecha', 'Total', 'Estado']]
    for orden in ordenes:
        data.append([
            f'OC-{orden.id}',
            orden.solicitante.username if orden.solicitante else 'N/A',
            orden.fecha_emision.strftime('%d/%m/%Y'),
            f'${orden.monto_total:,.0f}',
            orden.get_estado_display()
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@admin_required
def export_proveedores_excel(request):
    """Exporta la lista de proveedores a Excel."""
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Header
    headers = ['Nombre', 'RUT', 'Contacto', 'Teléfono', 'Email', 'Categoría', 'Tiempo Entrega', 'Monto Mínimo']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for proveedor in proveedores:
        ws.append([
            proveedor.nombre,
            proveedor.rut,
            proveedor.contacto,
            proveedor.telefono,
            proveedor.email,
            proveedor.categoria.nombre if proveedor.categoria else 'N/A',
            proveedor.tiempo_entrega,
            float(proveedor.monto_minimo)
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

@admin_required
def export_proveedores_pdf(request):
    """Exporta la lista de proveedores a PDF."""
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=proveedores.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Lista de Proveedores</b>", styles['Title'])
    elements.append(title)
    
    # Table data
    data = [['Nombre', 'RUT', 'Contacto', 'Teléfono', 'Email']]
    for proveedor in proveedores:
        data.append([
            proveedor.nombre,
            proveedor.rut,
            proveedor.contacto,
            proveedor.telefono,
            proveedor.email
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response
